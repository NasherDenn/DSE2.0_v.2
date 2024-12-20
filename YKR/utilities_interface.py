# -*- coding: utf-8 -*-

import datetime
import itertools
import time
import openpyxl
from PIL import Image

from YKR.utilities_db import *
from YKR.utilities_add_reports import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment


# проверка выбраны ли фильтры и введены ли данные для поиска
def not_check_data_and_filter(data_filter_for_search: dict, data_for_search: dict) -> bool:
    check_filter_location = False
    check_filter_method = False
    check_filter_year = False
    data = False
    for location in data_filter_for_search['location']:
        if data_filter_for_search['location'][location]:
            check_filter_location = True
    for method in data_filter_for_search['method']:
        if data_filter_for_search['method'][method]:
            check_filter_method = True
    for year in data_filter_for_search['year']:
        if data_filter_for_search['year'][year]:
            check_filter_year = True
    for key in data_for_search.keys():
        if data_for_search[key][0] != '':
            data = True
    if check_filter_location and check_filter_method and check_filter_year and data:
        return False
    else:
        return True


# названия БД по выбранным фильтрам в которых надо искать данные
def define_db_for_search(data_filter_for_search: dict) -> list:
    db_for_search = list()
    location = list()
    method = list()
    year = list()
    # выбираем выбранные фильтры
    for check_filter in data_filter_for_search['location']:
        if data_filter_for_search['location'][check_filter]:
            location.append(check_filter)
    for check_filter in data_filter_for_search['method']:
        if data_filter_for_search['method'][check_filter]:
            method.append(check_filter)
    for check_filter in data_filter_for_search['year']:
        if data_filter_for_search['year'][check_filter]:
            year.append(check_filter[2:])
    all_combinations = list(itertools.product(location, method, year))
    for combination in all_combinations:
        db_for_search.append(f'reports_db_{combination[0]}_{combination[1]}_{combination[2]}.sqlite')
    print(db_for_search)
    return db_for_search


# формируем кнопки названий найденных таблиц
def table_name_buttons(frame_for_table, y1, authorization, table, find_date, language, line_for_table_name_buttons):
    # задаём название кнопки по номеру репорта и помещаем внутрь frame
    button_for_table = QPushButton(second_underlining(table, find_date, language, line_for_table_name_buttons), frame_for_table)
    # координата отступа от левого края (меняется, когда происходит авторизация пользователя - появляется check box)
    if authorization:
        x1 = 20
    else:
        x1 = 0
    # задаём размеры и место расположения кнопки во frame
    button_for_table.setGeometry(QRect(x1, y1, 900, 20))
    # задаём стиль шрифта
    font_button_for_table = QFont()
    font_button_for_table.setFamily(u"Calibri")
    font_button_for_table.setPointSize(10)
    button_for_table.setStyleSheet('text-align: left; font: bold italic')
    button_for_table.setFont(font_button_for_table)
    button_for_table.show()
    # скрываем границы кнопки
    button_for_table.setFlat(True)
    # делаем кнопку переключателем
    button_for_table.setCheckable(True)
    return button_for_table


# формируем кнопки названий чертежей для таблиц
def drawing_name_buttons(frame_for_table, y1, x11, language, index_draw, path_drawing, draw):
    # задаём название кнопки по номеру репорта и помещаем внутрь frame
    drawing_button_for_table = QPushButton(second_underlining_drawing(language, index_draw), frame_for_table)
    # задаём размеры и место расположения кнопки во frame
    drawing_button_for_table.setGeometry(QRect(x11, y1, 90, 20))
    # задаём стиль шрифта
    font_drawing_button_for_table = QFont()
    font_drawing_button_for_table.setFamily(u"Calibri")
    font_drawing_button_for_table.setPointSize(10)
    drawing_button_for_table.setStyleSheet('text-align: left; font: bold italic')
    drawing_button_for_table.setFont(font_drawing_button_for_table)
    drawing_button_for_table.show()
    # скрываем границы кнопки
    drawing_button_for_table.setFlat(True)
    # открытие чертежа по нажатию на кнопку
    drawing_button_for_table.clicked.connect(lambda: Image.open(f'{path_drawing}\\{draw}').show())
    return drawing_button_for_table


def check_box_name_buttons(frame_for_table, y1, authorization):
    # задаём флажок для каждой кнопки номера репорта
    check_box = QCheckBox(frame_for_table)
    # задаём координаты флажка
    check_box.move(0, y1)
    if authorization:
        check_box.show()
    else:
        check_box.hide()
    return check_box


# формируем название кнопки для отображаемой таблицы
def second_underlining(table: str, find_date: str, language: str, line_for_table_name_buttons: str) -> str:
    number_report = table.replace('_', '-')[1:]
    date = find_date
    name_button = ''
    if language == 'ru':
        name_button = f'номер отчёта: {number_report}, дата: {date}, объект контроля: {line_for_table_name_buttons}'
    if language == 'en':
        name_button = f'report number: {number_report}, date: {date}, object of control: {line_for_table_name_buttons}'
    if language == 'kz':
        name_button = f'есеп нөмірі: {number_report}, күні: {date}, бақылау объектісі: {line_for_table_name_buttons}'
    return name_button


# кнопка названия чертежа
def second_underlining_drawing(language, index_draw):
    index_draw = int(index_draw) + 1
    name_button = ''
    if language == 'ru':
        name_button = f'чертёж-{index_draw}'
    if language == 'en':
        name_button = f'drawing-{index_draw}'
    if language == 'kz':
        name_button = f'есурет салу-{index_draw}'
    return name_button


# сортируем БД в порядке убывания
def sort_year(db_year: dict) -> list:
    unsort_db_year = []
    for db in db_year.keys():
        unsort_db_year.append(db)
    sort_db_year = sorted(unsort_db_year, reverse=True)
    return sort_db_year


# преобразование списка ключей (дат) в отсортированный список в формате iso-format
# find_data[0][db].keys() = list_date
def sort_date(list_date):
    # список сортированных дат
    sort_list_date = []
    # словарь форматированная дата: старая дата
    i = {}
    for index_date, date in enumerate(list_date):
        old_date = date
        if '-' in date:
            ind = date.index('-')
            date = date[ind + 1:]
        if re.search('[a-zA-Z]', date):
            sort_list_date.append(datetime.datetime.strptime(date, '%d.%B.%Y').isoformat()[:-9])
            i[datetime.datetime.strptime(date, '%d.%B.%Y').isoformat()[:-9]] = old_date
        else:
            try:
                sort_list_date.append(datetime.datetime.strptime(date, '%d.%m.%Y').isoformat()[:-9])
                i[datetime.datetime.strptime(date, '%d.%m.%Y').isoformat()[:-9]] = old_date
            except ValueError:
                logger_with_user.error(f'Ошибка в формате даты:\n'
                                       f'{date}\n'
                                       f'{traceback.format_exc()}')
    return i


# ширина фрейма для вывода найденных таблиц по умолчанию
new_width_frame = 1680


# функция отображения и повторного скрытия таблиц в frame
# l_t_v = list_table_view = список всех таблиц
# l_b_t = list_button_for_table = список всех номеров (кнопок) репортов
# l_ch_b = list_check_box = список всех чек-боксов
# y_1 - координата строки (кнопки) с номером репорта
# y_2 = y_1 + 20 - координата таблицы (20 - высота строки с номером репорта)
# l_h_t_v = list_height_table_view = список высот таблиц (строка с названием колонок + все строки таблицы)
# l_b_f_d = list_button_for_drawing = список кнопок с чертежами
# def visible_table_view(l_t_v, l_b_t, l_ch_b, l_h_t_v, authorization):
# list_drawing_button - список чертежей в папке для репорта
def visible_table_view(l_t_v, l_b_t, l_ch_b, l_h_t_v, authorization, l_b_f_d, frame_for_table, list_drawing_button):
    # x1 - координата кнопки таблицы
    if authorization:
        x1 = 20
    else:
        x1 = 0
    # y_1 - координата первой строки номера репорта
    y_1 = 0
    # y_2 - координата первой таблицы
    y_2 = 20
    ii = 0
    # список новых координат номеров репортов
    position_y1 = []
    # список новых координат таблиц
    position_y2 = []
    # обнуляем список нажатых кнопок таблиц
    list_button_for_table_true = []
    # обнуляем список нажатых кнопок чертежей
    list_button_for_drawing_true = []
    # список отжатых кнопок таблиц
    list_button_for_table_false = []
    # список отжатых кнопок чертежей
    list_button_for_drawing_false = []
    for i in l_b_t:
        # если нажата
        if i.isChecked():
            list_button_for_table_true.append(ii)
            list_button_for_drawing_true.append(ii)
        # если не нажата
        if not i.isChecked():
            list_button_for_table_false.append(ii)
            list_button_for_drawing_false.append(ii)
        ii += 1
    # вычисляем новые координаты номеров репортов и таблиц в зависимости от списка нажатых (list_button_for_table_true)
    # и не нажатых (list_button_for_table_false) кнопок
    for i in range(len(l_h_t_v)):
        # если нажата кнопка номера репорта
        if list_button_for_table_true:
            # перебираем номера нажатых кнопок
            for ii in list_button_for_table_true:
                if i == ii:
                    # добавляем в список координату кнопки номера репорта
                    position_y1.append(y_1)
                    # меняем координату кнопки номера репорта, потому что она нажата и появляется таблица с данными
                    y_1 += 40 + l_h_t_v[i]
                    # добавляем в список координату таблицы с данными
                    position_y2.append(y_2)
                    # меняем координату кнопки номера репорта, потому что она нажата и появляется таблица с данными
                    y_2 += 40 + l_h_t_v[ii]
        # если НЕ нажата кнопка номера репорта
        if list_button_for_table_false:
            # перебираем номера НЕ нажатых кнопок
            for ii in list_button_for_table_false:
                if i == ii:
                    # добавляем в список Не нажатой координату кнопки номера репорта
                    position_y1.append(y_1)
                    y_1 += 20
                    # добавляем в список координату таблицы с данными при НЕ нажатой кнопки номера репорта
                    position_y2.append(y_2)
                    y_2 += 20
    # общее количество строк в найденных таблицах
    all_count_row_in_search = 0
    # количество всех найденных таблиц
    all_count_table_in_search = len(l_h_t_v)
    frame_height_for_data_output = all_count_table_in_search * 20
    # делаем таблицы видимыми или скрываем их в зависимости от статуса
    for b in list_button_for_table_true:
        # передвигаем кнопку репорта
        l_b_t[b].move(x1, position_y1[b])
        # x11 - координата кнопки чертежей
        if authorization:
            x11 = 920
        else:
            x11 = 900
        # передвигаем кнопку чертежей
        for button_draw in l_b_f_d[b]:
            button_draw.move(x11, position_y1[b])
            x11 += 110
        l_t_v[b].setGeometry((QRect(0, position_y2[b], 1640, l_h_t_v[b])))
        # делаем таблицу из списка видимой
        l_t_v[b].setVisible(True)
        if l_ch_b[b]:
            # передвигаем флажок
            l_ch_b[b].move(0, position_y2[b] - 20)
        # количество строк в открытых таблицах
        all_count_row_in_search += l_h_t_v[b]
        # высота фрейма = количество таблиц * 20 (высоты строки с кнопкой для таблицы) + количество строк в открытых таблицах *
        # * 20 (высота строки в таблице) + количество открытых таблиц * 20 (расстояние между открытыми таблицами)
        frame_height_for_data_output = all_count_table_in_search * 20 + all_count_row_in_search + len(list_button_for_table_true) * 20
    for bb in list_button_for_table_false:
        # передвигаем кнопку репорта
        l_b_t[bb].move(x1, position_y1[bb])
        # x11 - координата кнопки чертежей
        if authorization:
            # для чертежй
            x11 = 920
        else:
            # для чертежй
            x11 = 900
        # передвигаем кнопку чертежей
        for button_draw in l_b_f_d[bb]:
            button_draw.move(x11, position_y1[bb])
            x11 += 110
        # делаем таблицу из списка снова скрытой
        l_t_v[bb].hide()
        if l_ch_b[bb]:
            # передвигаем флажок
            l_ch_b[bb].move(0, position_y1[bb])
    # если хоть в одном репорте больше чем 9 чертежей, то изменяем ширину поля для вывода найденных таблиц
    if len(list_drawing_button) > 7:
        global new_width_frame
        new_width_frame = 1680 + (len(list_drawing_button) - 7) * 110
    frame_for_table.setGeometry(0, 0, new_width_frame, frame_height_for_data_output)


# функция для вывода найденных открытых репортов и сводных данных на лист Excel для дальнейшей печати на принтер
def output_data_master(data_filter_for_search: list):
    wbb = openpyxl.Workbook()
    # дата и время формирования файла Excel для печати
    date_time_for_print = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    thin = Side(border_style="thin", color="000000")
    # создаём новый лист на каждую БД
    for db in data_filter_for_search:
        sheet_for_print = wbb.create_sheet(db)

        # вставляем в ячейку "A1" название столбца "Юнит"
        sheet_for_print.cell(row=1, column=1, value=str("Юнит"))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=1).font = Font(bold=True)
        # центрируем запись внутри
        sheet_for_print.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        # Устанавливаем ширину столбца "A"
        sheet_for_print.column_dimensions['A'].width = 20
        # выделяем её границами
        sheet_for_print.cell(row=1, column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # вставляем в ячейку "B1" название столбца "Номер репорта"
        sheet_for_print.cell(row=1, column=2, value=str(db))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=2).font = Font(bold=True)
        # центрируем запись внутри
        sheet_for_print.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')
        # Устанавливаем ширину столбца "B"
        sheet_for_print.column_dimensions['B'].width = 40
        # выделяем её границами
        sheet_for_print.cell(row=1, column=2).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # вставляем в ячейку "C1" название столбца "Дата репорта"
        sheet_for_print.cell(row=1, column=3, value=str("Дата репорта"))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=3).font = Font(bold=True)
        # центрируем запись внутри
        sheet_for_print.cell(row=1, column=3).alignment = Alignment(horizontal='center', vertical='center')
        # Устанавливаем ширину столбца "C"
        sheet_for_print.column_dimensions['C'].width = 15
        # выделяем её границами
        sheet_for_print.cell(row=1, column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # вставляем в ячейку "D1" название столбца "Work order"
        sheet_for_print.cell(row=1, column=4, value=str("Work order"))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=4).font = Font(bold=True)
        # центрируем запись внутри
        sheet_for_print.cell(row=1, column=4).alignment = Alignment(horizontal='center', vertical='center')
        # Устанавливаем ширину столбца "D"
        sheet_for_print.column_dimensions['D'].width = 15
        # выделяем её границами
        sheet_for_print.cell(row=1, column=4).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # вставляем в ячейку "E1" название столбца "Загружено таблиц в БД / всего таблиц в файле"
        sheet_for_print.cell(row=1, column=5, value=str("Загружено таблиц в БД / всего таблиц в файле"))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=5).font = Font(bold=True)
        # центрируем запись внутри и переносим по словам
        sheet_for_print.cell(row=1, column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Устанавливаем ширину столбца "E"
        sheet_for_print.column_dimensions['E'].width = 15
        # выделяем её границами
        sheet_for_print.cell(row=1, column=5).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # вставляем в ячейку "F1" название столбца "Список таблиц в БД"
        sheet_for_print.cell(row=1, column=6, value=str("Список таблиц в БД"))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=6).font = Font(bold=True)
        # центрируем запись внутри
        sheet_for_print.cell(row=1, column=6).alignment = Alignment(horizontal='center', vertical='center')
        # Устанавливаем ширину столбца "F"
        sheet_for_print.column_dimensions['F'].width = 45
        # выделяем её границами
        sheet_for_print.cell(row=1, column=6).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # закрепляем первую строку с названием столбцов
        sheet_for_print.freeze_panes = "A2"
        # перебираем строки в master
        for i, row in enumerate(data_master(db)):
            # перебираем столбцы в master
            for j, value in enumerate(row):
                # записываем значения в ячейки Excel
                sheet_for_print.cell(row=i + 2, column=j + 1, value=row[j])
                # центрируем запись внутри
                sheet_for_print.cell(row=i + 2, column=j + 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # выделяем её границами
                sheet_for_print.cell(row=i + 2, column=j + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # путь сохранения в папке с программой
        new_path_for_print_statistic = os.path.abspath(os.getcwd()) + '\\Print\\Statistic for print\\' + date_time_for_print[:7] + '\\'
        if not os.path.exists(new_path_for_print_statistic):
            # то создаём эту папку
            os.makedirs(new_path_for_print_statistic)
        # переменная имени файла с расширением для сохранения и последующего открытия
        name_for_print_statistic = f'{date_time_for_print} {db}.xlsx'
    # Удаление листа, создаваемого по умолчанию, при создании документа
    del wbb['Sheet']

    try:
        # сохраняем файл
        wbb.save(f'{new_path_for_print_statistic}{name_for_print_statistic}')
    except UnboundLocalError:
        pass
    else:
        wbb.close()
        # и открываем его
        os.startfile(new_path_for_print_statistic + name_for_print_statistic)


# достаём информацию о БД и номере(-ах) репорта(-ов) в котором(-ых) надо удалить таблицу(-ы)
def del_report(table_for_delete: list) -> dict:
    # словарь "БД": "список таблиц на удаление"
    db_table_for_delete = dict()
    # список таблиц для удаления
    list_table_for_delete_in_db = []
    # активатор первого определения таблиц и БД для удаления
    first_check_db = True
    # проверочная БД
    check_db = ''
    # количество БД в которых надо удалять таблицы
    count_table = len(table_for_delete)
    # перебираем выбранные названия таблиц для удаления
    for index_table, table in enumerate(table_for_delete):
        index_colon = table.text().index(':')
        index_comma = table.text().index(',')
        table = f'_{table.text()[index_colon + 2:index_comma]}'
        # если это первая таблица
        if first_check_db:
            # определяем БД
            db_for_delete = reports_db(table, True)[0]
            # обновляем проверочную БД
            check_db = db_for_delete
            # меняем активатор первого определения
            first_check_db = False
            # название таблицы для удаления из БД
        # если это не первая таблица
        else:
            db_for_delete = reports_db(table, True)[0]
        table = table.replace("-", "_")
        # если это вся та же БД
        if check_db == db_for_delete:
            # то продолжаем наполнять список таблиц в это БД для удаления
            list_table_for_delete_in_db.append(table)
            # если это ВООБЩЕ последняя таблица
            if index_table == count_table - 1:
                db_table_for_delete[db_for_delete] = list_table_for_delete_in_db
        # если же это НОВАЯ БД
        else:
            db_table_for_delete[check_db] = list_table_for_delete_in_db
            list_table_for_delete_in_db = []
            # если это ВООБЩЕ последняя таблица
            if index_table == count_table - 1:
                list_table_for_delete_in_db.append(table)
                db_table_for_delete[db_for_delete] = list_table_for_delete_in_db
            # иначе начинаем добавлять таблицы в новый список для следующей БД
            else:
                list_table_for_delete_in_db.append(table)
                check_db = db_for_delete
    return db_table_for_delete


# формируем название закладки листа Excel для печати
def name_table_for_excel_print(full_name_button: str) -> str:
    check_since = full_name_button.index(':') + 2
    check_to = full_name_button.index(',')
    name = full_name_button[check_since:check_to]
    return name


# заставка перед запуском программы
def splash_screen():
    splash_pix = QPixmap(f'{os.path.abspath(os.getcwd())}\\Images\\splash_screen.png')
    splash = QSplashScreen(splash_pix, Qt.WindowStaysOnTopHint)
    opaqueness = 0.0
    step = 0.005
    splash.setWindowOpacity(opaqueness)
    splash.show()
    while opaqueness < 2.3:
        splash.setWindowOpacity(opaqueness)
        time.sleep(step)
        opaqueness += step
    splash.close()


# создаём файл Excel после начала верификации
# def create_print_verification():
#     wbb = openpyxl.Workbook()
#     # дата и время формирования файла Excel для печати
#     date_time_for_print = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
#     # путь сохранения в папке с программой
#     new_path_report_verification = os.path.abspath(os.getcwd()) + '\\Print\\Report of verification\\' + date_time_for_print[:7] + '\\'
#     # переменная имени файла с расширением для сохранения и последующего открытия
#     name_report_verification = str(date_time_for_print) + ' Report of verification' + '.xlsx'
#     if not os.path.exists(new_path_report_verification):
#         # то создаём эту папку
#         os.makedirs(new_path_report_verification)
#     # сохраняем файл
#     wbb.save(new_path_report_verification + name_report_verification)
#     wbb.close()
#     return new_path_report_verification, name_report_verification


# вывод на лист Excel результатов верификации
# args[0] - путь к файлу Excel
# args[1] - имя файла Excel
# args[2] - название закладки по выбранной опции верификации
# def print_verification(*args):
#     wb = openpyxl.load_workbook(args[0] + args[1])
#     # создаём новый лист, если его нет, на каждую выбранную опцию верификации
#     if args[2] not in wb.sheetnames:
#         sheet_for_print = wb.create_sheet(args[2])
#     # вставляем в первую строку заголовка
#     sheet_for_print.cell(row=1, column=1, value=str('Всего загружено'))
#     sheet_for_print.cell(row=1, column=2, value=str('Таблицы'))






    # wb.save(args[0] + args[1])
    # thin = Side(border_style="thin", color="000000")
    #
    # # название листа для таблицы
    # name_table = name_table_for_excel_print( push_button.text())
    # # создаём новый лист на каждую таблицу
    # sheet_for_print = wbb.create_sheet(name_table)
    # # вставляем в первую строку название кнопки по выбранной таблицу
    # sheet_for_print.cell(row=1, column=1, value=str(push_button.text()))
#         #         # выделяем её жирным
#         #         sheet_for_print.cell(row=1, column=1).font = Font(bold=True)
#         #         # объединяем в первой строке столбцы 'A:J'
#         #         sheet_for_print.merge_cells('A1:J1')
#         #         # вставляем во вторую строку названия столбцов
#         #         # перебираем названия столбцов
#         #         for index_column in range(full_sqm[index_push_button].columnCount()):
#         #             sheet_for_print.cell(row=2, column=index_column + 1,
#         #                                  value=str(full_sqm[index_push_button].query().record().fieldName(index_column)))
#         #             # выделяем её жирным
#         #             sheet_for_print.cell(row=2, column=index_column + 1).font = Font(bold=True)
#         #             # центрируем запись внутри
#         #             sheet_for_print.cell(row=2, column=index_column + 1).alignment = Alignment(horizontal='center', vertical='center')
#         #             # заполняем лист Excel
#         #             for index_row in range(full_sqm[index_push_button].rowCount()):
#         #                 sheet_for_print.cell(row=index_row + 3, column=index_column + 1,
#         #                                      value=str(full_sqm[index_push_button].record(index_row).value(index_column)))
#         #                 # выделяем основные данные границами
#         #                 sheet_for_print.cell(row=index_row + 3, column=index_column + 1).border = Border(top=thin, left=thin, right=thin,
#         #                                                                                                  bottom=thin)
#         #             # закрепляем первую строку с названием кнопки, по которой выбрана таблица, и вторую с названиями столбцов
#         #             sheet_for_print.freeze_panes = "A3"
#         #             # выделяем её границами
#         #             sheet_for_print.cell(row=2, column=index_column + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
#         #         # ручной автоподбор ширины столбцов по содержимому
#         #         ascii_range = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
#         #                        'S', 'T', 'V', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI',
#         #                        'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AV', 'AX', 'AY', 'AZ']
#         #         # перебираем все заполненные столбцы
#         #         for collumn in range(1, full_sqm[index_push_button].columnCount() + 1):
#         #             max_length_column = 0
#         #             # перебираем все заполненные строки
#         #             for roww in range(2, full_sqm[index_push_button].rowCount() + 1):
#         #                 if len(str(sheet_for_print.cell(row=roww, column=collumn).value)) > max_length_column:
#         #                     max_length_column = len(str(sheet_for_print.cell(row=roww, column=collumn).value))
#         #             # устанавливаем ширину заполненных столбцов по их содержимому
#         #             sheet_for_print.column_dimensions[ascii_range[collumn]].width = max_length_column + 2
#         #     # путь сохранения в папке с программой
#         #     new_path_for_print = os.path.abspath(os.getcwd()) + '\\Print\\Report for print\\' + date_time_for_print[:7] + '\\'
#         #     if not os.path.exists(new_path_for_print):
#         #         # то создаём эту папку
#         #         os.makedirs(new_path_for_print)
#         #     # переменная имени файла с расширением для сохранения и последующего открытия
#         #     name_for_print = str(date_time_for_print) + ' Report for print' + '.xlsx'
#         # # Удаление листа, создаваемого по умолчанию, при создании документа
#         # del wbb['Sheet']
#         # # сохраняем файл
#         # wbb.save(new_path_for_print + name_for_print)
#         # wbb.close()
#         # # и открываем его
#         # os.startfile(new_path_for_print + name_for_print)
